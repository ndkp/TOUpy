##############################
#プログラムの解説(1)

import openpyxl as excel, json
#入出力ファイル名
in_file = "matome.json"
out_dir = "./invoice04"
template_file = "invoice-template.xlsx"
subject = "4月分のご請求"

#メイン処理
def gen_invoice():
    #JSONファイルを読み込む
    with open(in_file, "rt") as fp:
        customers = json.load(fp)
    #顧客ごとに請求書を作成
    for name, data in customers.items():
        make_invoice(name, data)

##############################
#プログラムの解説(2)

#テンプレートにデータを流し込む
def make_invoice(name, data):
    #テンプレートを読み込む
    book = excel.load_workbook(template_file)
    sheet = book.active
    #シートに名前と件名と請求金額を書き込む
    sheet["B4"] = name
    sheet["C10"] = subject
    sheet["C11"] = data["total"]
    #明細データを連続で書き込む
    for i, values in enumerate(data["items"]):
        (date, item, price, cnt) = values
        row  = 14 + i
        sheet.cell(row, 2, date)      #B列
        sheet.cell(row, 3, item)      #C列
        sheet.cell(row, 5, price)     #E列
        sheet.cell(row, 6, cnt)       #F列
        sheet.cell(row, 7, price*cnt) #G列
############################
#プログラムの解説(3)
    #請求書をファイルに保存
    out_file = out_dir + "/" + name + " 様.xlsx"
    book.save(out_file)
    print("save: ", out_file)

#メインプログラムを実行
if __name__ == "__main__":
    gen_invoice() #メイン処理を実行
