import openpyxl as excel

book = excel.load_workbook(
    "uriage.xlsx", data_only=True)
sheet = book.active

it = sheet.iter_rows(
    min_row=3, min_col=1)

for row in it:
    values = [cell.value for cell in row]
    name = values[6]
    if name not in book.sheetnames:
        to_sheet = book.create_sheet(title=name)
        to_sheet["A1"] = f"{name}の売上一覧"
        to_sheet.append(['NO', '納品日', '商品名',
            '単価', '数量', '金額', '担当営業'])
    else:
        to_sheet = book[name]
    to_sheet.append(values)
    mrow = to_sheet.max_row
    to_sheet.cell(mrow, 2).number_format = "m月d日"

book.save("uriage_split.xlsx")
