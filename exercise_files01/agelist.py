import openpyxl as excel
import datetime

book = excel.Workbook()
sheet = book.active

thisyear = datetime.date.today().year

for i in range(101):
    age = i
    year = thisyear - age
    age_cell = sheet.cell(i+1, 1)
    age_cell.value = f"{age}歳"
    year_cell = sheet.cell(i+1, 2)
    year_cell.value = f"{year}年"

book.save("agelist.xlsx")
