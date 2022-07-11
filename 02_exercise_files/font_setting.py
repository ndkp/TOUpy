import openpyxl as excel
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, PatternFill

book = excel.Workbook()
sheet = book.active
sheet.column_dimensions['A'].width = 25

cell = sheet["A1"]
cell.value = "フォント設定練習"

cell.alignment = Alignment(
    horizontal = 'center', vertical = 'center')

side = Side(style='thin', color='000000')
cell.border = Border(
    left=side, right=side, top=side, bottom=side)

cell.font = Font(
    name = 'メイリオ', size = 14,
    bold = True, color = 'FF0000')

cell.fill = PatternFill(
    patternType = 'solid',
    fgColor = 'D9D9D9')

book.save("font_setting.xlsx")
